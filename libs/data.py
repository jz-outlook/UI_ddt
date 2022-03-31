import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def find_none(l):
    """
    返回列表中的非None项
    :param l:
    :return:
    """
    new_l = []
    for _ in l:
        if _ is not None:
            new_l.append(_)
    return new_l


def case_by_excel(path):
    """
    读取excel文件内容并返回测试用例内容
    :param path: 文件的路径
    :return:
    """

    wb = load_workbook(path)  # 整个工作簿（整个文件）

    for ws in wb.worksheets:  # 遍历excel中的sheet页
        # 在遍历sheet页时，创建新的套件
        _suite = {
            "info": {
                "name": ws.title,  # sheet页的名称
            },
            "cases": {},  # 套件中的用例
        }
        _cases = _suite["cases"]

        new_case_id = 0  # 给用例设置ID

        for (step_id, step_name, keyword, *args,) in ws.iter_rows(min_row=2, values_only=True):  # 遍历sheet中的行
            if step_id < 0:  # case中的info部分
                new_case_id += 1  # 新的测试用例
                # 根据id获取用例，如果没有就创建一个新的
                new_case = _cases.get(new_case_id, {'info': {}, 'steps': []})
                new_case['info'][keyword] = args[0]
            else:  # 用例的步骤部分
                new_case = _cases.get(new_case_id, {'info': {}, 'steps': []})
                # [_ for _ in args if _] 如果参数是None就忽略
                # new_case['steps'].append((keyword, [_ for _ in args if _]))
                new_case['steps'].append((step_id, step_name, keyword, find_none(args)))

            # print(f'步骤id{step_id}', f'步骤名称{step_name}', f'关键字{keyword}', f'参数{args}', '---' * 10)

            _cases[new_case_id] = new_case
        yield _suite
